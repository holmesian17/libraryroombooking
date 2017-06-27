import openpyxl, sys
wb = openpyxl.load_workbook('Library Room Booking.xlsx')

def ask():    
  while True:   
    print('When is the booking date? (mm/dd/yyyy)')
    date = input()
    sheet.cell(row=rownum, column=3).value = date
    print('What time is the booking? (24hrs - 4 digits)')
    later_time = input()
    sheet.cell(row=rownum, column=4).value = later_time
    print('What is the borrower barcode?')
    barcode = input()
    print(barcode)
    print('Is this correct? (Y or N)')
    verify = input()
    verify = verify.lower()
    while True:
      if verify.startswith('n'):
        continue
      else:
        sheet.cell(row=rownum, column=1).value = barcode
        print('How many people will be using the room?') #TODO: if persons < 2, may not book???
        persons = input()
        sheet.cell(row=rownum, column=5).value = persons
        equip_initials()
        wb.save('Library Room Booking.xlsx')
        sys.exit("Done")
        break

def equip_initials():
  print('Do they need equipment?')
  equip_need = input()
  equip_need = equip_need.lower()
  if equip_need.startswith('n'):    
    print("Are there any notes you'd like to add?") 
    noteQuestion = input()
    noteQuestion = noteQuestion.lower()
    if noteQuestion.startswith('y'):
      print('Enter your note:')
      note = input()
      sheet.cell(row=rownum, column=6).value = note
      print("Enter the staff member's initials")
      initials = input()
      initials = initials.lower()
      sheet.cell(row=rownum, column=7).value = initials
    else:
      sheet.cell(row=rownum, column=6).value= 'N/A'  
      print("Enter the staff member's initials")
      initials = input()
      initials = initials.lower()
      sheet.cell(row=rownum, column=7).value = initials
  elif equip_need.startswith('y'):
    print('What equipment?')
    equip = input()
    sheet.cell(row=rownum, column=6).value = equip
    print("Are there any notes you'd like to add?")
    noteQuestion = input()
    noteQuestion = noteQuestion.lower()
    if noteQuestion.startswith('y'):
      print('Enter your note:')
      note = input()
      sheet.cell(row=rownum, column=6).value = note
      print("Enter the staff member's initials")
      initials = input()
      initials = initials.lower()
      sheet.cell(row=rownum, column=7).value = initials
      while True:
        break
    else:
      sheet.cell(row=rownum, column=6).value = 'N/A'
      print("Enter the staff member's initials")
      initials = input()
      initials = initials.lower()
      sheet.cell(row=rownum, column=7).value = initials
      while True:
        break

while True:
  print('What room is being booked? (P, S)')
  room = input()
  room = room.lower()
  if room == 'p':
    sheet = wb.get_sheet_by_name('Pye')
    sheet['B2'] = 'Pye'
  elif room == 's':
    sheet = wb.get_sheet_by_name('Scriver')
    sheet['B2'] = 'Scriver'
  else:
    continue 
  rownum = sheet.max_row + 1
  ask()




#TODO: make it fill the excel sheet in the next open row
#TODO: need to record the date and time the booking is taken
#TODO: need the dates and times to be accurate - if not in certain format, then redo??? or would that cause problems?
#TODO: need the thing to kick back in spots where they entered a wrong thing i.e. not what they're supposed to
#TODO: have it automatically add it into outlook meeting room calendar
#TODO: get it to run as an exe on computers at work

#have an alert come up 2 hours after their booking to have the staff person check with them
#making an alarm that will go off after their 2 hours is up - asks if wanting to continue
#if B then go through a different set of steps or questions for Bunday room?
  #would this then automatically get sent to Teresa?
