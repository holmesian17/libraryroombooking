import openpyxl, sys
wb = openpyxl.load_workbook('Library Room Booking.xlsx')

def date():     
    print('What is the booking date? (mm/dd/yyyy)')
    datebooked = input()
    sheet.cell(row=rownum, column=3).value = datebooked
    print('What is the date the room is needed? If today, press \'T\' ')
    dateneeded = input()
    dateneeded = dateneeded.lower()
    if dateneeded == 't':
      sheet.cell(row=rownum, column=4).value = datebooked
    else:
      sheet.cell(row=rownum, column=4).value = dateneeded
    print('What time is the booking? (24hrs - 4 digits)')
    later_time = input()
    sheet.cell(row=rownum, column=5).value = later_time

def patronName():
  print('What is the patron\'s name? (lastname, firstname)')
  name = input()
  sheet.cell(row=rownum, column=1).value = name

def barcodeask():
  print('What is the borrower barcode?')
  barcode = input()
  print(barcode)
  print('Is this correct? (Y or N)')
  verify = input()
  verify = verify.lower()
  if verify.startswith('y'):
    sheet.cell(row=rownum, column=2).value = barcode
    print('How many people will be using the room?')
    persons = input()
    sheet.cell(row=rownum, column=6).value = persons
  else:
    barcodeask() 

def equip_initials():
  print('Do they need equipment?')
  equip_need = input()
  equip_need = equip_need.lower()
  if equip_need.startswith('n'):  
    sheet.cell(row=rownum, column=8).value = 'N/A'  
  elif equip_need.startswith('y'):
    print('What equipment?')
    equip = input()
    sheet.cell(row=rownum, column=8).value = equip
  else:
    print('That is not a valid response.')
    equip_initials()

def noteQuestion():
  print("Are there any notes you'd like to add?") 
  noteQuestiontext = input()
  noteQuestiontext = noteQuestiontext.lower()
  if noteQuestiontext.startswith('y'):
    print('Enter your note:')
    note = input()
    sheet.cell(row=rownum, column=7).value = note
    print("Enter the staff member's initials")
    initials = input()
    initials = initials.lower()
    sheet.cell(row=rownum, column=9).value = initials
    while True:
      break
  elif noteQuestiontext.startswith('n'):
    sheet.cell(row=rownum, column=7).value= 'N/A'  
    print("Enter the staff member's initials")
    initials = input()  
    initials = initials.lower()
    sheet.cell(row=rownum, column=9).value = initials 
    while True:
      break
  else:
    print('That is not a valid response.')
    noteQuestion()    

while True:
  print('What room is being booked? (P, S)')
  room = input()
  room = room.lower()
  if room == 'p':
    sheet = wb.get_sheet_by_name('Pye')
    sheet['B2'] = 'Pye'
  elif room == 's':
    sheet = wb.get_sheet_by_name('Scriver')
    sheet['B2'] = 'Scriver'
  else:
    continue
  rownum = sheet.max_row + 1
  date()
  patronName()
  barcodeask()
  equip_initials()
  noteQuestion()
  wb.save('Library Room Booking.xlsx')
  sys.exit("Done")



#TODO: need to record the date and time the booking is taken
#TODO: need the dates and times to be accurate - if not in certain format, then redo??? or would that cause problems?
#TODO: have it automatically add it into outlook meeting room calendar
#TODO: get it to run as an exe on computers at work
#TODO: enter their name for the outlook event - use pyexchange  

#do we want a name spot? contact info? - Yes
#have an alert come up 2 hours after their booking to have the staff person check with them
#making an alarm that will go off after their 2 hours is up - asks if wanting to continue
#if B then go through a different set of steps or questions for Bunday room?
  #would this then automatically get sent to Teresa?
